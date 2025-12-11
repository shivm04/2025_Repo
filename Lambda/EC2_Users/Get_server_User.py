import boto3
import time
import json
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ssm = boto3.client("ssm")
ses = boto3.client("ses")

LOGO_URL = "https://example.com/"


# ---------------------------------------------------------
# Run SSM Command (exclude root, include only /bin/bash users)
# ---------------------------------------------------------
def run_ssm_command(instance_id):
    cmd = """
users=$(awk -F: '$7=="/bin/bash"{print $1":"$6":"$7}' /etc/passwd)

for line in $users; do
    username=$(echo $line | cut -d: -f1)
    home=$(echo $line | cut -d: -f2)
    shell=$(echo $line | cut -d: -f3)

    if [ "$username" = "root" ]; then
        continue
    fi

    PRIV=$(groups $username | grep -E "(sudo|wheel)" -q && echo "sudo" || echo "normal")
    MFA=$( [ -f $home/.google_authenticator ] && echo "Yes" || echo "No" )

    echo "$username,$PRIV,$MFA,$home,$shell"
done
"""
    response = ssm.send_command(
        InstanceIds=[instance_id],
        DocumentName="AWS-RunShellScript",
        Parameters={"commands": [cmd]},
    )

    cmd_id = response["Command"]["CommandId"]
    time.sleep(3)

    output = ssm.get_command_invocation(
        CommandId=cmd_id, InstanceId=instance_id
    )

    return output["StandardOutputContent"]


# -----------------------------------
# Excel Styles
# -----------------------------------
def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def style_cell(cell):
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    cell.alignment = Alignment(horizontal="left")


# -----------------------------------
# Excel Generator
# -----------------------------------
def generate_excel(data):
    wb = Workbook()
    wb.remove(wb.active)

    for server in data:
        sheet = wb.create_sheet(title=server["instance_name"])

        # Bold top headers
        sheet["A1"] = "Instance Name:"
        sheet["A1"].font = Font(bold=True, size=12)

        sheet["B1"] = server["instance_name"]
        sheet["B1"].font = Font(size=12)

        sheet["A2"] = "Instance ID:"
        sheet["A2"].font = Font(bold=True, size=12)

        sheet["B2"] = server["instance_id"]
        sheet["B2"].font = Font(size=12)

        sheet.append([])

        # Table header
        sheet.append(["Sr No", "Username", "Privilege", "2FA", "Home Dir", "Shell"])

        # Apply header style
        for col in range(1, 7):
            style_header(sheet.cell(row=4, column=col))

        # Data rows
        for i, row in enumerate(server["users"], start=1):
            sheet.append([
                i,
                row["username"],
                row["priv"],
                row["mfa"],
                row["home"],
                row["shell"]
            ])

        # Style data rows
        for row in sheet.iter_rows(min_row=5, max_col=6):
            for cell in row:
                style_cell(cell)

        # Auto column width
        for col in sheet.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[letter].width = max_len + 4

    file = "/tmp/server_users.xlsx"
    wb.save(file)
    return file


# -----------------------------------
# Email sender
# -----------------------------------
def send_email(from_email, to_email, summary, attachment_path):

    with open(attachment_path, "rb") as f:
        file_data = f.read()

    attachment_b64 = base64.b64encode(file_data).decode()

    # Summary table
    summary_html = """
    <table style="border-collapse: collapse; width: 80%; margin: auto; font-family: Arial;">
        <tr style="background-color:#4F81BD; color:white;">
            <th style="padding:8px;">Server Name</th>
            <th style="padding:8px;">Instance ID</th>
            <th style="padding:8px;">Total Users</th>
        </tr>
    """

    for item in summary:
        summary_html += f"""
        <tr style="background:#f2f2f2;">
            <td style="padding:8px; text-align:center;">{item['instance_name']}</td>
            <td style="padding:8px; text-align:center;">{item['instance_id']}</td>
            <td style="padding:8px; text-align:center;">{item['total_users']}</td>
        </tr>
        """

    summary_html += "</table>"

    # Email HTML
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; background-color:#fafafa; padding:20px;">

        <div style="text-align:center;">
            <img src="{LOGO_URL}" width="140" style="margin-bottom:20px;"/>
        </div>

        <h2 style="text-align:center; color:#333;">Phicommerce - Server User Report</h2>
        <p style="text-align:center; color:#555;">Below is the summary of all servers:</p>

        {summary_html}

        <p style="text-align:center; margin-top:25px; color:#555;">
            The detailed Excel report is attached.
        </p>

    </body>
    </html>
    """

    # MIME message
    msg = f"""From: {from_email}
To: {to_email}
Subject: Phicommerce - Server Users Report
MIME-Version: 1.0
Content-Type: multipart/mixed; boundary="NextPart"

--NextPart
Content-Type: text/html

{html_body}

--NextPart
Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; name="server_users.xlsx"
Content-Transfer-Encoding: base64
Content-Disposition: attachment; filename="server_users.xlsx"

{attachment_b64}
--NextPart--
"""

    ses.send_raw_email(
        Source=from_email,
        Destinations=[to_email],
        RawMessage={"Data": msg}
    )


# -----------------------------------
# Main Lambda Handler
# -----------------------------------
def lambda_handler(event, context):

    from_email = event.get("from_email")
    to_email = event.get("to_email")
    instances = event.get("instances", [])

    if not from_email or not to_email:
        return {"error": "from_email and to_email required in test event"}

    results = []
    summary = []

    for inst in instances:
        instance_id = inst["InstanceId"]
        instance_name = inst["InstanceName"]

        output = run_ssm_command(instance_id)
        rows = []

        for line in output.splitlines():
            try:
                username, priv, mfa, home, shell = line.split(",")
                rows.append({
                    "username": username,
                    "priv": priv,
                    "mfa": mfa,
                    "home": home,
                    "shell": shell
                })
            except:
                pass

        results.append({
            "instance_id": instance_id,
            "instance_name": instance_name,
            "users": rows
        })

        summary.append({
            "instance_name": instance_name,
            "instance_id": instance_id,
            "total_users": len(rows)
        })

    excel_path = generate_excel(results)

    send_email(from_email, to_email, summary, excel_path)

    return {
        "status": "success",
        "message": "Email sent with improved Excel and professional HTML template."
    }
